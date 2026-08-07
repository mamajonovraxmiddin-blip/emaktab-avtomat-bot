import os
import io
import openpyxl
import asyncio
import sqlite3
from aiogram import Bot, Dispatcher, types, F
from aiogram.filters import Command
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.utils.keyboard import InlineKeyboardBuilder
from emaktab import run_emaktab_login
from dotenv import load_dotenv
from cryptography.fernet import Fernet  # Shifrlash uchun

load_dotenv()

BOT_TOKEN = os.getenv("BOT_TOKEN")
SPECIAL_CODE = os.getenv("SPECIAL_CODE", "12345")
ENCRYPTION_KEY = os.getenv("ENCRYPTION_KEY")

# Shifrlash menejerini ishga tushirish
fernet = Fernet(ENCRYPTION_KEY.encode())

bot = Bot(token=BOT_TOKEN)
dp = Dispatcher()

# ----------------- MA'LUMOTLAR BAZASI BILAN ISHLASH (SQLITE) -----------------
DB_NAME = "emaktab_bot.db"

def init_db():
    """Ma'lumotlar bazasi va jadvallarni yaratish"""
    conn = sqlite3.connect(DB_NAME)
    cursor = conn.cursor()
    # Sinf nomini saqlash uchun jadval
    cursor.execute('''CREATE TABLE IF NOT EXISTS config (key TEXT PRIMARY KEY, value TEXT)''')
    # Foydalanuvchilar (O'quvchi va ota-onalar) jadvali
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_type TEXT,
            name TEXT,
            login TEXT,
            password TEXT
        )
    ''')
    conn.commit()
    conn.close()

def set_class_name(class_name):
    conn = sqlite3.connect(DB_NAME)
    cursor = conn.cursor()
    cursor.execute("INSERT OR REPLACE INTO config (key, value) VALUES ('class_name', ?)", (class_name,))
    conn.commit()
    conn.close()

def get_class_name():
    conn = sqlite3.connect(DB_NAME)
    cursor = conn.cursor()
    cursor.execute("SELECT value FROM config WHERE key = 'class_name'")
    row = cursor.fetchone()
    conn.close()
    return row[0] if row else "Noma'lum"

def save_users_to_db(user_type, users_list):
    """Foydalanuvchilarni parolini shifrlab bazaga saqlash"""
    conn = sqlite3.connect(DB_NAME)
    cursor = conn.cursor()
    # Avvalgi eski ro'yxatni o'chirib tashlaymiz (Refresh funksiyasi uchun)
    cursor.execute("DELETE FROM users WHERE user_type = ?", (user_type,))
    
    for user in users_list:
        # Parolni xavfsiz AES-256 ko'rinishida shifrlash
        encrypted_password = fernet.encrypt(user['password'].encode()).decode()
        cursor.execute(
            "INSERT INTO users (user_type, name, login, password) VALUES (?, ?, ?, ?)",
            (user_type, user['name'], user['login'], encrypted_password)
        )
    conn.commit()
    conn.close()

def get_users_from_db(user_type):
    """Bazadan foydalanuvchilarni o'qish va parolini deshifrlash"""
    conn = sqlite3.connect(DB_NAME)
    cursor = conn.cursor()
    cursor.execute("SELECT name, login, password FROM users WHERE user_type = ?", (user_type,))
    rows = cursor.fetchall()
    conn.close()
    
    users_list = []
    for row in rows:
        try:
            # Shifrlangan parolni qayta asl holiga keltirish
            decrypted_password = fernet.decrypt(row[2].encode()).decode()
            users_list.append({
                "name": row[0],
                "login": row[1],
                "password": decrypted_password
            })
        except Exception:
            continue
    return users_list

# Bazani dastur boshlanishida faollashtiramiz
init_db()
# -----------------------------------------------------------------------------

class BotStates(StatesGroup):
    waiting_for_class = State()
    waiting_for_code = State()
    main_menu = State()

def create_excel_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Ismi va Familiyasi", "Login", "Parol"])
    ws.append(["Eshmatov Toshmat", "eshmat123", "pas12345"])
    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)
    return file_stream

def build_dynamic_keyboard(user_type):
    builder = InlineKeyboardBuilder()
    # Ma'lumotlarni RAMdan emas, xavfsiz bazadan yuklaymiz
    data_list = get_users_from_db(user_type)
    
    for idx, user in enumerate(data_list):
        builder.button(text=f"👤 {user['name']}", callback_data=f"login:{user_type}:{idx}")
    builder.adjust(2)
    
    builder.row(types.InlineKeyboardButton(text="🔄 Ro'yhatni yangilash", callback_data=f"refresh:{user_type}"))
    builder.row(types.InlineKeyboardButton(text="⬅️ Ortga", callback_data="back_to_menu"))
    return builder.as_markup()

@dp.message(Command("start"))
async def cmd_start(message: types.Message, state: FSMContext):
    await message.answer("✨ **Salom, men emaktab.uz saytiga kiradigan botman.**", parse_mode="Markdown")
    await message.answer("📝 Iltimos sinfingiz raqami va nomini kriting:\n*(Misol uchun: 9-A, 4-B)*", parse_mode="Markdown")
    await state.set_state(BotStates.waiting_for_class)

@dp.message(BotStates.waiting_for_class, F.text)
async def process_class(message: types.Message, state: FSMContext):
    set_class_name(message.text.strip()) # Bazaga saqlash
    await message.answer("🔒 **Iltimos maxsus kodni kriting:**", parse_mode="Markdown")
    await state.set_state(BotStates.waiting_for_code)

@dp.message(BotStates.waiting_for_code, F.text)
async def process_code(message: types.Message, state: FSMContext):
    user_code = message.text.strip()
    try:
        await message.delete()
    except Exception:
        pass
        
    if user_code == SPECIAL_CODE:
        await message.answer("🔒 Maxsus kod: `*****` *(Yashirildi)*", parse_mode="Markdown")
        
        kb = [
            [types.KeyboardButton(text="👨‍🎓 O'quvchilar ro'yhati")],
            [types.KeyboardButton(text="👨‍👩‍👦 Ota-onalar ro'yhati")]
        ]
        keyboard = types.ReplyKeyboardMarkup(keyboard=kb, resize_keyboard=True)
        class_name = get_class_name()
        await message.answer(f"✅ Kod to'g'ri! Tizim tayyor. Sinf: **{class_name}**\nQuyidagi bo'limlardan birini tanlang:", reply_markup=keyboard, parse_mode="Markdown")
        await state.set_state(BotStates.main_menu)
    else:
        await message.answer("❌ Maxsus kod noto'g'ri! Qayta kriting:")

@dp.message(BotStates.main_menu, F.text.in_(["👨‍🎓 O'quvchilar ro'yhati", "👨‍👩‍👦 Ota-onalar ro'yhati"]))
async def process_menu_selection(message: types.Message):
    user_type = "students" if "O'quvchilar" in message.text else "parents"
    current_list = get_users_from_db(user_type)
    
    if not current_list:
        excel_file = create_excel_template()
        file_input = types.BufferedInputFile(excel_file.read(), filename=f"{user_type}_namuna.xlsx")
        await message.reply_document(
            file_input, 
            caption=f"📋 Sizda hali ro'yxat yuklanmagan.\nUshbu toza namuna faylini yuklab oling, to'ldirib botga qayta yuboring."
        )
    else:
        markup = build_dynamic_keyboard(user_type)
        await message.answer(f"📋 **{message.text}** bo'limi foydalanuvchini tanlang:", reply_markup=markup, parse_mode="Markdown")

@dp.message(BotStates.main_menu, F.document)
async def handle_excel_upload(message: types.Message):
    document = message.document
    if not document.file_name.endswith(('.xlsx', '.xls')):
        await message.answer("⚠️ Iltimos, faqat Excel (.xlsx) formatidagi faylni jo'nating.")
        return

    status_msg = await message.answer("📥 Fayl o'qilmoqda va xavfsiz shifrlanmoqda, iltimos kuting...")
    file_bytes = io.BytesIO()
    await bot.download(document, destination=file_bytes)
    file_bytes.seek(0)
    
    try:
        wb = openpyxl.load_workbook(file_bytes)
        ws = wb.active
        
        parsed_data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] and row[1] and row[2]:
                parsed_data.append({
                    "name": str(row[0]).strip(),
                    "login": str(row[1]).strip(),
                    "password": str(row[2]).strip()
                })
        
        user_type = "parents" if "parent" in document.file_name.lower() else "students"
        
        # Ma'lumotlarni bazaga xavfsiz shifrlab yozish
        save_users_to_db(user_type, parsed_data)
        
        markup = build_dynamic_keyboard(user_type)
        await status_msg.delete()
        await message.answer(f"✨ Ro'yxat muvaffaqiyatli yuklandi va ma'lumotlar bazasida shifrlandi! Jami: **{len(parsed_data)}** ta xavfsiz tugma hosil bo'ldi.", reply_markup=markup, parse_mode="Markdown")
        
    except Exception as e:
        await status_msg.edit_text(f"❌ Faylni qayta ishlashda xatolik yuz berdi: {e}")

@dp.callback_query(F.data.startswith("login:"))
async def handle_silent_login(callback: types.CallbackQuery):
    _, user_type, idx = callback.data.split(":")
    idx = int(idx)
    
    data_list = get_users_from_db(user_type)
    user = data_list[idx]
    
    report_msg = await callback.message.answer(f"⏳ **{user['name']}** uchun eMaktab tizimiga orqa fonda jimgina kirilmoqda. Iltimos kuting...", parse_mode="Markdown")
    await callback.answer()
    
    result = await run_emaktab_login(user['login'], user['password'])
    
    if result["status"]:
        await report_msg.edit_text(f"✅ **Hisobot Holati:**\n👤 Foydalanuvchi: {user['name']}\n📊 Natija: {result['message']}", parse_mode="Markdown")
    else:
        await report_msg.edit_text(f"❌ **Hisobot Holati:**\n👤 Foydalanuvchi: {user['name']}\n📊 Xatolik sababi: `{result['message']}`", parse_mode="Markdown")

@dp.callback_query(F.data.startswith("refresh:"))
async def handle_refresh(callback: types.CallbackQuery):
    user_type = callback.data.split(":")[1]
    excel_file = create_excel_template()
    file_input = types.BufferedInputFile(excel_file.read(), filename=f"{user_type}_namuna.xlsx")
    await callback.message.reply_document(file_input, caption="🔄 Yangi ro'yxat namuna fayli. To'ldirib qayta yuboring.")
    await callback.answer()

@dp.callback_query(F.data == "back_to_menu")
async def handle_back(callback: types.CallbackQuery):
    """Ortga tugmasi bosilganda asosiy menyu matnini ko'rsatish"""
    await callback.message.answer("🏡 Asosiy menyudasiz. Pastdagi menyu tugmalaridan foydalaning.")
    await callback.answer()

if __name__ == '__main__':
    asyncio.run(dp.start_polling(bot))

